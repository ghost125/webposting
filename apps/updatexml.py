import base64
import os
from io import BytesIO
import time
import datetime

import streamlit as st 
import pandas as pd
import openpyxl


from .postxml_ema02 import prettyxmlfile, postxml_ema



def xmldownload(xmlfile):
    buffered = BytesIO()
    file = open(xmlfile,'rb').read()
    buffered.write(file)
    b64 = base64.b64encode(buffered.getvalue()).decode()
    href = f'<a href="data:file/xml;base64,{b64}" download="result.xml">Download XML file</a>'
    return href


def add_block(checklist,label,inexcel,showlevel=4):
    #checklist.append('NONE')
    st.write(f'## Section of {label}')
    dfindex = len(checklist)-1
    temp = st.selectbox(label,checklist,index=dfindex)
    groups = 'test'
    if temp != 'NONE':
        df = pd.read_excel(inexcel,temp,dtype='object')
        if showlevel==4:
            groups = st.multiselect("Groups",df.columns,key='g_'+temp)
            total = st.multiselect('Total',df.columns,key='t_'+temp)
            sets = st.multiselect('Analysis Sets',df.columns,key='s_'+temp)
        elif showlevel == 3:
            groups = st.multiselect("Groups",df.columns,key='g_'+temp)
            sets = st.multiselect('Analysis Sets',df.columns,key='s_'+temp)
        elif showlevel == 2:
            groups = st.multiselect("Groups",df.columns,key='g_'+temp)
        else:
            pass
        st.write('The data is:')
        st.dataframe(df)
        if showlevel==4:
            return df, groups,total, sets
        elif showlevel==3:
            return df, groups,sets
        elif showlevel==2:
            return df, groups
        else:
            return df
    else:
        return None


def app():
    inxml = st.file_uploader('Upload XML file',key='Upload_xml')
    incountry = st.file_uploader('Upload Eucat contry excel file',key='Upload_country',type='XLSX')
    if incountry:
        country_sheets = openpyxl.load_workbook(incountry).sheetnames
        country=st.selectbox("Country:",country_sheets)
    else:
        incountry='euctid_country.xlsx'
        country='country'
    
    country_sheets = openpyxl.load_workbook(incountry).sheetnames
    country=st.selectbox("Country:",country_sheets)
    eucat_id=pd.read_excel(incountry,country)
    st.write('The Country EuCAT ID is:')
    st.dataframe(eucat_id)
 
    inexcel = st.file_uploader('Upload Excel created base on TLF meta',key='Upload_excel',type='XLSX')
    if inexcel:
        insheets = openpyxl.load_workbook(inexcel).sheetnames
        if isinstance(insheets,str):
            insheets = [insheets]
        if 'NONE' not in insheets:
            insheets=insheets+['NONE']
        country_input = add_block(insheets,'Country',inexcel,1)
        age_input = add_block(insheets,'Age',inexcel,2)
        disp_input = add_block(insheets,'Disposition',inexcel,2)
        age_n_input = add_block(insheets,'Age continue',inexcel,3)
        age_c_input = add_block(insheets,'Age category',inexcel)
        gender_input = add_block(insheets,'Gender',inexcel)
        demo_c_input = add_block(insheets,'Demography category',inexcel)
        demo_n_input = add_block(insheets,'Demography continue',inexcel,3)
        endpoints_input = add_block(insheets,'Endpoints',inexcel,3)
        stats_input = add_block(insheets,'Stat. analysis',inexcel,1)
        if st.button('Update XML'):
            test = postxml_ema(inxml)
            if len(country_input)>0:
                test.add_country_count(country_input,incountry,country)
            if age_input:
                test.add_agegroup(age_input[0],age_input[1][0])
            if disp_input:
                test.add_patientflow(disp_input[0],disp_input[1])
            if gender_input:
                test.add_gender_cat_characters(gender_input[0],gender_input[1],gender_input[2],
                gender_input[3])
            if age_n_input:
                test.add_age_cont_characters(age_n_input[0],age_n_input[1],age_n_input[2])
            if age_c_input:
                test.add_age_cat_characters(age_c_input[0],age_c_input[1],age_c_input[2],
                age_c_input[3])
            if demo_c_input:
                test.add_baseline_cat_characters(demo_c_input[0],demo_c_input[1],demo_c_input[2],
                demo_c_input[3])
            if demo_n_input:
                test.add_baseline_cont_characters(demo_n_input[0],demo_n_input[1],demo_n_input[2])
            if endpoints_input:
                test.add_endpoints(endpoints_input[0],endpoints_input[1],endpoints_input[2],
                stats_input)
            filename = 'sp'+datetime.datetime.now().strftime("%Y%m%d%h%m%c").replace(':','').replace(' ','')+'.xml'
            test.tofile(filename,pretty=False)
            prettyxmlfile(filename)
            st.markdown(xmldownload(filename),unsafe_allow_html=True)
            os.remove(filename)
    
